# -*- coding: utf-8 -*-
"""
对公账户开户智能体 - 后端服务入口
基于 AgentScope 2.0 + FastAPI 构建
"""
import os
import sys
import json
import uuid
import asyncio
from datetime import datetime
from typing import Optional

import uvicorn
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from dotenv import load_dotenv

from agentscope.agent import Agent
from agentscope.model import OpenAIChatModel
from agentscope.credential import OpenAICredential
from agentscope.message import UserMsg
from agentscope.event import EventType

from agent_config import AGENT_CONFIG
from auth import (
    UserRegister,
    UserLogin,
    UserInfo,
    TokenResponse,
    user_store,
    create_token,
    get_current_user,
)

# 导入 Coordinator 和技能（触发注册）
from coordinator import route_intent
from skills import skill_registry
import skills.potential_customer  # noqa: F401 — 触发技能注册
import skills.risk_check  # noqa: F401 — 触发风险预查技能注册
import skills.customer_outreach  # noqa: F401 — 触发拓户准备技能注册
import skills.product_recommend  # noqa: F401 — 触发产品智荐技能注册
import skills.product_match  # noqa: F401 — 触发产品智能匹配技能注册
import skills.account_opening  # noqa: F401 — 触发对公账户开户技能注册
from context_memory import context_memory
from follow_up_agent import predict_follow_up

# 加载环境变量
load_dotenv()

# ============================================================
# FastAPI 应用初始化
# ============================================================
app = FastAPI(
    title="对公账户开户智能体 API",
    description="基于 AgentScope 2.0 的对公账户开户智能助手",
    version="1.0.0",
)

# CORS 中间件 - 支持前后端分离
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载 H5 静态页面
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(STATIC_DIR):
    app.mount("/h5", StaticFiles(directory=STATIC_DIR + "/h5"), name="h5")


# ============================================================
# 数据模型
# ============================================================
class ChatRequest(BaseModel):
    message: str
    conversation_id: Optional[str] = None


class ConversationCreate(BaseModel):
    title: Optional[str] = "新对话"


class Message(BaseModel):
    id: str
    role: str  # "user" | "assistant"
    content: str
    created_at: str


class Conversation(BaseModel):
    id: str
    user_id: str
    title: str
    messages: list[Message]
    created_at: str
    updated_at: str


class ConversationListItem(BaseModel):
    id: str
    user_id: str
    title: str
    message_count: int
    created_at: str
    updated_at: str


# ============================================================
# 内存存储 - 会话管理（按用户隔离）
# ============================================================
# 结构: { user_id: { conv_id: Conversation } }
conversations: dict[str, dict[str, Conversation]] = {}

# 追踪每个会话中已调用过的技能（按企业分组，用于追问agent反循环）
# 结构: { conversation_id: { credit_code: [skill_name, ...] } }
conversation_skills: dict[str, dict[str, list[str]]] = {}

# 开户提交通知队列（H5 页面提交后回调写入，前端轮询读取后清除）
# 结构: { conversation_id: [notification_dict, ...] }
account_notifications: dict[str, list[dict]] = {}


def _get_user_convs(user_id: str) -> dict[str, Conversation]:
    """获取指定用户的会话字典"""
    if user_id not in conversations:
        conversations[user_id] = {}
    return conversations[user_id]


def create_conversation(user_id: str, title: str = "新对话") -> Conversation:
    """创建新会话"""
    conv_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    conv = Conversation(
        id=conv_id,
        user_id=user_id,
        title=title,
        messages=[],
        created_at=now,
        updated_at=now,
    )
    _get_user_convs(user_id)[conv_id] = conv
    return conv


# ============================================================
# AgentScope Agent 实例管理
# ============================================================
def create_agent() -> Agent:
    """创建对公账户开户智能体实例"""
    api_key = os.getenv("DEEPSEEK_API_KEY")
    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com")

    if not api_key:
        print("=" * 60)
        print("警告: 未设置 DEEPSEEK_API_KEY 环境变量")
        print("请在 backend/.env 文件中配置:")
        print("  DEEPSEEK_API_KEY=你的DeepSeek API密钥")
        print("  DEEPSEEK_BASE_URL=https://api.deepseek.com  (可选,默认值)")
        print("获取 API Key: https://platform.deepseek.com/api_keys")
        print("=" * 60)

    return Agent(
        name=AGENT_CONFIG["name"],
        system_prompt=AGENT_CONFIG["system_prompt"],
        model=OpenAIChatModel(
            credential=OpenAICredential(
                api_key=api_key or "",
                base_url=base_url,
            ),
            model=AGENT_CONFIG["model_name"],
        ),
    )


# 全局 Agent 实例（延迟初始化，首次调用时创建）
_agent: Optional[Agent] = None


def get_agent() -> Agent:
    """获取或创建 Agent 实例"""
    global _agent
    if _agent is None:
        _agent = create_agent()
    return _agent


# ============================================================
# API 路由
# ============================================================

# -------------------------------------------------------------------
# 认证接口（无需登录）
# -------------------------------------------------------------------
@app.post("/api/auth/register", response_model=TokenResponse)
async def register(body: UserRegister):
    """用户注册"""
    user = user_store.create_user(body.username, body.password)
    if not user:
        raise HTTPException(status_code=409, detail="用户名已存在")
    token = create_token(user["id"], user["username"])
    return TokenResponse(
        access_token=token,
        user=UserInfo(**user),
    )


@app.post("/api/auth/login", response_model=TokenResponse)
async def login(body: UserLogin):
    """用户登录"""
    user = user_store.authenticate(body.username, body.password)
    if not user:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_token(user["id"], user["username"])
    return TokenResponse(
        access_token=token,
        user=UserInfo(**user),
    )


@app.get("/api/user/me", response_model=UserInfo)
async def get_me(current_user: UserInfo = Depends(get_current_user)):
    """获取当前登录用户信息（验证 Token 有效性）"""
    return current_user


# -------------------------------------------------------------------
# 健康检查（无需登录）
# -------------------------------------------------------------------
@app.get("/api/health")
async def health_check():
    """健康检查接口"""
    return {
        "status": "ok",
        "service": "对公账户开户智能体",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat(),
    }


# -------------------------------------------------------------------
# 会话管理接口（需登录）
# -------------------------------------------------------------------
@app.get("/api/conversations")
async def list_conversations(current_user: UserInfo = Depends(get_current_user)):
    """获取当前用户的会话列表"""
    result = []
    for conv in _get_user_convs(current_user.id).values():
        result.append({
            "id": conv.id,
            "user_id": conv.user_id,
            "title": conv.title,
            "message_count": len(conv.messages),
            "created_at": conv.created_at,
            "updated_at": conv.updated_at,
        })
    result.sort(key=lambda x: x["updated_at"], reverse=True)
    return {"conversations": result}


@app.post("/api/conversations")
async def new_conversation(
    body: ConversationCreate,
    current_user: UserInfo = Depends(get_current_user),
):
    """创建新会话"""
    conv = create_conversation(current_user.id, title=body.title or "新对话")
    return {
        "id": conv.id,
        "user_id": conv.user_id,
        "title": conv.title,
        "created_at": conv.created_at,
    }


@app.get("/api/conversations/{conversation_id}")
async def get_conversation(
    conversation_id: str,
    current_user: UserInfo = Depends(get_current_user),
):
    """获取指定会话详情"""
    conv = _get_user_convs(current_user.id).get(conversation_id)
    if not conv:
        raise HTTPException(status_code=404, detail="会话不存在")
    return conv.model_dump()


@app.delete("/api/conversations/{conversation_id}")
async def delete_conversation(
    conversation_id: str,
    current_user: UserInfo = Depends(get_current_user),
):
    """删除指定会话"""
    user_convs = _get_user_convs(current_user.id)
    if conversation_id not in user_convs:
        raise HTTPException(status_code=404, detail="会话不存在")
    del user_convs[conversation_id]
    return {"status": "deleted", "id": conversation_id}


# -------------------------------------------------------------------
# 聊天接口（需登录）
# -------------------------------------------------------------------
@app.post("/api/chat/stream")
async def chat_stream(
    body: ChatRequest,
    current_user: UserInfo = Depends(get_current_user),
):
    """
    流式聊天接口 (SSE)
    发送用户消息，返回 Agent 的流式响应
    """
    user_convs = _get_user_convs(current_user.id)

    # 获取或创建会话
    conversation_id = body.conversation_id
    if not conversation_id or conversation_id not in user_convs:
        conv = create_conversation(current_user.id, title="新对话")
        conversation_id = conv.id
    else:
        conv = user_convs[conversation_id]

    user_msg_id = str(uuid.uuid4())
    now = datetime.now().isoformat()

    # 存储用户消息
    user_msg = Message(
        id=user_msg_id,
        role="user",
        content=body.message,
        created_at=now,
    )
    conv.messages.append(user_msg)
    conv.updated_at = now

    # ── 会话首次对话：记录创建时间为首条消息时间，标题为首条消息内容 ──
    if len(conv.messages) == 1:
        conv.created_at = now
        conv.title = body.message[:30] + ("..." if len(body.message) > 30 else "")

    async def event_generator():
        try:
            assistant_msg_id = str(uuid.uuid4())

            # 发送元数据
            yield f"data: {json.dumps({'type': 'meta', 'conversation_id': conversation_id})}\n\n"

            # ── 第1步：Coordinator 意图识别 ──
            decision = await route_intent(body.message)

            if decision.get("action") == "skill":
                # ── 第2步：执行技能 ──
                skill_name = decision.get("skill", "")
                skill_params = decision.get("params", {})

                # ── 上下文记忆：填充缺失的企业主体参数 ──
                ctx = context_memory.get(conversation_id)
                if not ctx.is_empty():
                    if not skill_params.get("company_name") and not skill_params.get("credit_code"):
                        if ctx.credit_code:
                            skill_params["credit_code"] = ctx.credit_code
                            if ctx.company_name:
                                skill_params["company_name"] = ctx.company_name
                            print(f"[Context] 自动填充 credit_code: {ctx.credit_code}, company_name: {ctx.company_name}")
                        elif ctx.company_name:
                            skill_params["company_name"] = ctx.company_name
                            print(f"[Context] 自动填充 company_name: {ctx.company_name}")

                print(f"[Chat] Coordinator 路由到技能: {skill_name}, 参数: {skill_params}")

                # 传递 conversation_id 给技能（供 H5 回调通知使用）
                skill_params["_conversation_id"] = conversation_id

                result = await skill_registry.invoke(
                    skill_name, current_user.id, skill_params
                )

                if result.get("error"):
                    # 技能执行异常，作为普通文本返回
                    yield f"data: {json.dumps({'type': 'text_delta', 'content': result['error'], 'message_id': assistant_msg_id})}\n\n"
                    yield f"data: {json.dumps({'type': 'text_done', 'content': result['error'], 'message_id': assistant_msg_id})}\n\n"
                else:
                    action = result.get("action", "")
                    if action == "summary":
                        yield f"data: {json.dumps({'type': 'potential_customer_summary', 'data': result, 'message_id': assistant_msg_id})}\n\n"
                    elif action == "detail":
                        yield f"data: {json.dumps({'type': 'potential_customer_detail', 'data': result, 'message_id': assistant_msg_id})}\n\n"
                    elif action in ("result", "ambiguous", "not_found"):
                        # 根据技能名区分事件类型
                        if skill_name == "prepare_customer_outreach":
                            event_type = "outreach_result"
                        elif skill_name == "recommend_products":
                            event_type = "product_recommend_result"
                        elif skill_name == "match_products_intelligently":
                            event_type = "product_match_result"
                        elif skill_name == "open_corporate_account":
                            event_type = "account_opening_result"
                        else:
                            event_type = "risk_check_result"
                        yield f"data: {json.dumps({'type': event_type, 'data': result, 'message_id': assistant_msg_id})}\n\n"

                # ── 上下文记忆：成功后更新当前企业主体 ──
                if action == "result" and result.get("credit_code"):
                    context_memory.update(
                        conversation_id,
                        company_name=result.get("company_name", ""),
                        credit_code=result.get("credit_code", ""),
                    )
                    print(f"[Context] 已更新上下文 → {result.get('company_name')} ({result.get('credit_code')})")

                # 存储结构化回复到会话
                summary_text = json.dumps(result, ensure_ascii=False)
                conv.messages.append(Message(
                    id=assistant_msg_id,
                    role="assistant",
                    content=summary_text,
                    created_at=datetime.now().isoformat(),
                ))
                conv.updated_at = datetime.now().isoformat()

                # ── 第3步：追踪已调用技能（按企业分组） + 调用追问agent ──
                if action == "result":
                    credit = result.get("credit_code", "")
                    conv_companies = conversation_skills.setdefault(conversation_id, {})
                    company_skills = conv_companies.setdefault(credit, []) if credit else []
                    if skill_name not in company_skills:
                        company_skills.append(skill_name)

                    # 汇总全会话所有企业的技能（供追问agent参考主体切换）
                    all_skills: list[str] = []
                    for code, skills in conv_companies.items():
                        for s in skills:
                            all_skills.append(s)

                    follow_up_text = await predict_follow_up(
                        skill_name=skill_name,
                        skill_action=action,
                        company_name=result.get("company_name", ""),
                        credit_code=result.get("credit_code", ""),
                        conversation_skills=list(all_skills),
                        current_company_skills=list(company_skills),
                    )
                    if follow_up_text:
                        yield f"data: {json.dumps({'type': 'follow_up_suggestion', 'content': follow_up_text, 'message_id': assistant_msg_id})}\n\n"

            else:
                # ── 普通聊天：走 AgentScope Agent ──
                agent = get_agent()
                full_content = ""

                async for evt in agent.reply_stream(
                    UserMsg("用户", body.message)
                ):
                    match evt.type:
                        case EventType.TEXT_BLOCK_START:
                            full_content = ""
                            yield f"data: {json.dumps({'type': 'text_start', 'message_id': assistant_msg_id})}\n\n"
                        case EventType.TEXT_BLOCK_DELTA:
                            if evt.delta:
                                full_content += evt.delta
                                yield f"data: {json.dumps({'type': 'text_delta', 'content': evt.delta, 'message_id': assistant_msg_id})}\n\n"
                        case EventType.TEXT_BLOCK_END:
                            yield f"data: {json.dumps({'type': 'text_done', 'content': full_content, 'message_id': assistant_msg_id})}\n\n"
                        case EventType.REPLY_START:
                            pass
                        case _:
                            pass

                # 存储助手消息
                if full_content:
                    assistant_msg = Message(
                        id=assistant_msg_id,
                        role="assistant",
                        content=full_content,
                        created_at=datetime.now().isoformat(),
                    )
                    conv.messages.append(assistant_msg)
                    conv.updated_at = assistant_msg.created_at

            # ── 更新会话标题：首次对话时用第一条用户消息作为标题 ──
            # （注：已在存用户消息时立即设置 created_at + title，此处保留作为兜底）
            if conv.title == "新对话" and len(conv.messages) >= 2:
                first_user_msg = ""
                for m in conv.messages:
                    if m.role == "user":
                        first_user_msg = m.content
                        break
                if first_user_msg:
                    conv.title = first_user_msg[:30] + ("..." if len(first_user_msg) > 30 else "")

            yield f"data: {json.dumps({'type': 'done', 'conversation_id': conversation_id})}\n\n"

        except Exception as e:
            error_msg = f"处理请求时发生错误: {str(e)}"
            print(f"[ERROR] {error_msg}", file=sys.stderr)
            yield f"data: {json.dumps({'type': 'error', 'content': error_msg})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ============================================================
# 风险预查 - H5 详细报告 API
# ============================================================
@app.get("/api/risk-report/{credit_code}")
async def get_risk_report(credit_code: str):
    """返回指定企业的完整风险报告数据（供H5页面使用）"""
    import json as _json
    from pathlib import Path as _Path

    risk_file = _Path(__file__).parent / "data" / "risk_check.json"
    if not risk_file.exists():
        raise HTTPException(status_code=404, detail="风险数据文件不存在")

    with open(risk_file, "r", encoding="utf-8") as f:
        risk_data = _json.load(f)

    result = risk_data.get(credit_code)
    if not result:
        raise HTTPException(status_code=404, detail=f"未找到信用代码 {credit_code} 的风险信息")

    return result


# ============================================================
# 拓户准备 - H5 谈资/话术 API
# ============================================================
@app.get("/api/outreach/{credit_code}")
async def get_outreach_data(credit_code: str):
    """返回指定企业的拓户准备完整数据（供H5页面使用）"""
    import json as _json
    from pathlib import Path as _Path

    outreach_file = _Path(__file__).parent / "data" / "customer_outreach.json"
    if not outreach_file.exists():
        raise HTTPException(status_code=404, detail="拓户准备数据文件不存在")

    with open(outreach_file, "r", encoding="utf-8") as f:
        outreach_data = _json.load(f)

    result = outreach_data.get(credit_code)
    if not result:
        raise HTTPException(status_code=404, detail=f"未找到信用代码 {credit_code} 的拓户准备数据")

    return result


# ============================================================
# 产品智荐 - H5 详细报告 API
# ============================================================
@app.get("/api/product-recommend/{credit_code}")
async def get_product_recommend(credit_code: str):
    """返回指定企业的产品推荐完整数据（供H5页面使用）"""
    import json as _json
    from pathlib import Path as _Path

    prod_file = _Path(__file__).parent / "data" / "product_recommendations.json"
    if not prod_file.exists():
        raise HTTPException(status_code=404, detail="产品推荐数据文件不存在")

    with open(prod_file, "r", encoding="utf-8") as f:
        prod_data = _json.load(f)

    recs = prod_data.get("recommendations", {})
    result = recs.get(credit_code)
    if not result:
        raise HTTPException(status_code=404, detail=f"未找到信用代码 {credit_code} 的产品推荐数据")

    # 组装返回给 H5 的数据
    product_pool = prod_data.get("products", {})
    sorted_recs = sorted(result["recommendations"], key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(r["priority"], 99))
    products = []
    for r in sorted_recs:
        prod = product_pool.get(r["key"], {})
        products.append({
            "product_name": prod.get("product_name", r["key"]),
            "category": prod.get("category", ""),
            "priority": r["priority"],
            "priority_label": {"high": "高优先级", "medium": "中优先级", "low": "低优先级"}.get(r["priority"], ""),
            "reason": r["reason"],
            "expected_amount": r.get("expected_amount", ""),
            "features": prod.get("features", []),
            "application_period": prod.get("application_period", ""),
        })

    return {
        "credit_code": result["credit_code"],
        "company_name": result["company_name"],
        "analysis_summary": result.get("analysis_summary", ""),
        "products": products,
        "total_count": len(products),
    }


# ============================================================
# 潜客清单 - 用户自定义上传 API
# ============================================================

@app.get("/api/customer-template")
async def get_customer_template(
    current_user: UserInfo = Depends(get_current_user),
):
    """下载客户清单上传模板（Excel .xlsx 格式）"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "客户清单"

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    headers = ["企业名称", "统一社会信用代码", "推荐得分(0-100)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写入示例数据
    sample_data = [
        ["示例科技有限公司", "91110108MA01B3XK2P", 85.5],
        ["示例供应链管理有限公司", "91440300MA5DCTJH8N", 78.0],
    ]
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="center")
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # 设置列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入内存并返回
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_template.xlsx"},
    )


@app.post("/api/customer-upload")
async def upload_customer_list(
    file: UploadFile = File(...),
    mode: str = Form("overwrite"),
    current_user: UserInfo = Depends(get_current_user),
):
    """
    上传自定义客户清单（Excel .xlsx 文件）
    mode: overwrite=覆盖存量数据, append=追加更新
    """
    import json as _json
    from pathlib import Path as _Path
    from openpyxl import load_workbook

    # 验证文件类型
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的 Excel 文件")

    # 解析 Excel
    try:
        content = await file.read()
        import io
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
        new_customers = []
        for row in rows:
            if not row or not row[0]:  # 跳过空行
                continue
            name = str(row[0]).strip()
            credit_code = str(row[1]).strip() if row[1] else ""
            score = float(row[2]) if row[2] else 0.0
            if name and credit_code:
                new_customers.append({
                    "name": name,
                    "credit_code": credit_code,
                    "score": score,
                })

        wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {str(e)}")

    if not new_customers:
        raise HTTPException(status_code=400, detail="Excel 文件中没有有效的客户数据")

    # 存储到 JSON 文件
    upload_file = _Path(__file__).parent / "data" / "uploaded_customers.json"
    if upload_file.exists():
        with open(upload_file, "r", encoding="utf-8") as f:
            all_data = _json.load(f)
    else:
        all_data = {}

    user_id = current_user.id

    if mode == "append":
        # 追加模式：合并去重（以 credit_code 为唯一键）
        existing = all_data.get(user_id, [])
        existing_codes = {c.get("credit_code") for c in existing}
        for c in new_customers:
            if c["credit_code"] not in existing_codes:
                existing.append(c)
            else:
                for i, ec in enumerate(existing):
                    if ec.get("credit_code") == c["credit_code"]:
                        existing[i] = c
                        break
        all_data[user_id] = existing
    else:
        # 覆盖模式
        all_data[user_id] = new_customers

    with open(upload_file, "w", encoding="utf-8") as f:
        _json.dump(all_data, f, ensure_ascii=False, indent=2)

    total = len(all_data[user_id])
    print(f"[Upload] 用户 {user_id} 上传客户清单(Excel), mode={mode}, 共 {total} 条")

    return {
        "status": "ok",
        "mode": mode,
        "total_count": total,
        "message": f"{'追加' if mode == 'append' else '覆盖'}上传成功，当前共 {total} 条客户记录",
    }


# ============================================================
# 对公账户开户 API 接口
# ============================================================

_ACCOUNT_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "account_opening")
_ACCOUNT_DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "account_opening.json")


def _load_account_apps() -> dict:
    if os.path.exists(_ACCOUNT_DATA_FILE):
        with open(_ACCOUNT_DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"applications": {}}


def _save_account_apps(data: dict):
    os.makedirs(os.path.dirname(_ACCOUNT_DATA_FILE), exist_ok=True)
    with open(_ACCOUNT_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.post("/api/account-opening/upload")
async def account_opening_upload(
    app_id: str = Form(...),
    business_license: Optional[UploadFile] = File(None),
    legal_rep_id: Optional[UploadFile] = File(None),
):
    """上传开户资料图片（营业执照/法人身份证）"""
    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")
    if app["status"] != "upload":
        raise HTTPException(status_code=400, detail=f"当前状态为{app['status']}，无法上传")

    # 保存文件
    os.makedirs(os.path.join(_ACCOUNT_UPLOAD_DIR, app_id), exist_ok=True)
    saved = {}
    for field, file_obj in [("business_license", business_license), ("legal_rep_id", legal_rep_id)]:
        if file_obj and file_obj.filename:
            ext = os.path.splitext(file_obj.filename)[1] or ".jpg"
            save_path = os.path.join(_ACCOUNT_UPLOAD_DIR, app_id, f"{field}{ext}")
            content = await file_obj.read()
            with open(save_path, "wb") as f:
                f.write(content)
            saved[field] = save_path
            print(f"[AccountOpening] 保存 {field} -> {save_path}")

    if not saved:
        raise HTTPException(status_code=400, detail="请至少上传一份资料")

    app["documents"] = saved
    app["status"] = "processing"
    _save_account_apps(apps)

    return {"status": "ok", "app_id": app_id, "next_step": "processing", "message": f"资料上传成功，共 {len(saved)} 份文件"}


@app.post("/api/account-opening/process/{app_id}")
async def account_opening_process(
    app_id: str,
):
    """触发 Mock OCR + 数据预填"""
    from skills.account_opening import _mock_ocr_and_prefill

    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")
    if app["status"] != "processing":
        raise HTTPException(status_code=400, detail=f"当前状态为{app['status']}，无法处理")

    # Mock OCR + 预填
    form_data = _mock_ocr_and_prefill(app["credit_code"], app["company_name"])
    app["form_data"] = form_data
    app["status"] = "preview"
    _save_account_apps(apps)

    base_url = os.getenv("BASE_URL", f"http://localhost:{os.getenv('PORT', '8000')}")
    conv_id = app.get("conversation_id", "")
    from urllib.parse import quote as _quote
    preview_url = f"{base_url}/h5/account-preview.html?app_id={app_id}&conversation_id={_quote(conv_id)}"

    print(f"[AccountOpening] 预填完成: {app['company_name']} ({app['credit_code']})")
    return {"status": "ok", "app_id": app_id, "preview_url": preview_url, "message": "数据预填完成，请预览确认"}


@app.get("/api/account-opening/preview/{app_id}")
async def account_opening_preview(
    app_id: str,
):
    """获取预填表单数据"""
    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")
    if app["status"] not in ("preview", "submitted"):
        raise HTTPException(status_code=400, detail=f"当前状态为{app['status']}，无法预览")

    return {
        "status": "ok",
        "app_id": app_id,
        "app_status": app["status"],
        "company_name": app["company_name"],
        "credit_code": app["credit_code"],
        "form_data": app["form_data"],
    }


@app.put("/api/account-opening/update/{app_id}")
async def account_opening_update(
    app_id: str,
    request: Request,
):
    """客户经理微调表单数据"""
    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")
    if app["status"] != "preview":
        raise HTTPException(status_code=400, detail=f"当前状态为{app['status']}，已提交后不可修改")

    body = await request.json()
    form_data = body.get("form_data")
    if not form_data:
        raise HTTPException(status_code=400, detail="请提供 form_data")

    app["form_data"] = form_data
    _save_account_apps(apps)

    return {"status": "ok", "app_id": app_id, "message": "表单数据已更新"}


@app.post("/api/account-opening/submit/{app_id}")
async def account_opening_submit(
    app_id: str,
):
    """提交确认，状态锁定"""
    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")
    if app["status"] == "submitted":
        raise HTTPException(status_code=400, detail="该申请已提交，不可重复提交")
    if app["status"] != "preview":
        raise HTTPException(status_code=400, detail=f"当前状态为{app['status']}，请先完成预览")

    app["status"] = "submitted"
    app["submitted_at"] = datetime.now().isoformat()
    _save_account_apps(apps)

    base_url = os.getenv("BASE_URL", f"http://localhost:{os.getenv('PORT', '8000')}")
    submitted_url = f"{base_url}/h5/account-submitted.html?app_id={app_id}"

    print(f"[AccountOpening] 已提交: {app['company_name']} ({app['credit_code']})")
    return {
        "status": "ok",
        "app_id": app_id,
        "submitted_url": submitted_url,
        "submitted_at": app["submitted_at"],
        "company_name": app["company_name"],
        "message": f"开户申请已成功提交！企业名称：{app['company_name']}",
    }


@app.post("/api/account-opening/notify/{app_id}")
async def account_opening_notify(
    app_id: str,
    request: Request,
):
    """
    H5 页面提交成功后的回调接口
    将提交通知写入队列，供前端轮询读取
    """
    apps = _load_account_apps()
    app = apps["applications"].get(app_id)
    if not app:
        raise HTTPException(status_code=404, detail="开户申请不存在")

    body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    conversation_id = body.get("conversation_id") or app.get("conversation_id", "")

    if not conversation_id:
        return {"status": "ok", "message": "无关联会话，跳过通知"}

    base_url = os.getenv("BASE_URL", f"http://localhost:{os.getenv('PORT', '8000')}")
    submitted_url = f"{base_url}/h5/account-submitted.html?app_id={app_id}"

    notification = {
        "type": "account_submitted",
        "app_id": app_id,
        "company_name": app.get("company_name", ""),
        "credit_code": app.get("credit_code", ""),
        "submitted_url": submitted_url,
        "submitted_at": app.get("submitted_at", datetime.now().isoformat()),
    }

    queue = account_notifications.setdefault(conversation_id, [])
    queue.append(notification)
    print(f"[AccountOpening] 回调通知已入队: conv={conversation_id}, app={app_id}, company={app.get('company_name')}")

    return {"status": "ok", "message": "通知已入队"}


@app.get("/api/account-opening/notifications/{conversation_id}")
async def get_account_notifications(
    conversation_id: str,
):
    """
    查询并清除指定会话的待处理开户通知（前端轮询调用）
    """
    notifications = account_notifications.pop(conversation_id, [])
    return {"notifications": notifications}


# ============================================================
# 应用启动
# ============================================================
if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    print(f"启动对公账户开户智能体服务...")
    print(f"API 文档: http://localhost:{port}/docs")
    print(f"健康检查: http://localhost:{port}/api/health")
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=port,
        reload=True,
    )
